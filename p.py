import openpyxl

# Load the Excel sheets
workbook_a = openpyxl.load_workbook('sheet_a.xlsx')
workbook_b = openpyxl.load_workbook('sheet_b.xlsx')

# Select the first sheet in each workbook
sheet_a = workbook_a.worksheets[0]
sheet_b = workbook_b.worksheets[0]

# Iterate over all cells in sheet A
list = []
for row in sheet_a.iter_rows():
    for cell in row:

        # Get the cell value as a string
        cell_value = str(cell.value)

        # Get the last 7 digits of the cell value
        search_value = cell_value[-7:]

        # Search for the search value in sheet B
        for row_b in sheet_b.iter_rows():
            for cell_b in row_b:

                # Get the cell value as a string
                cell_value_b = str(cell_b.value)

                # Check if the search value is a substring of the cell value
                if search_value in cell_value_b and search_value != 'None':
                    list.append({cell_value, "found in sheet B:", cell_value_b})
print(list)

# Print a message to confirm that the script has finished running
print('Search completed successfully.')
